from PIL import Image
from splinter import Browser
from resolver import quebrar_captcha
import time
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
import openpyxl 
import sklearn
import logging
logging.basicConfig(level=logging.WARNING)
import logging
import sys








# Verificar se a pasta "ler" existe, caso contrário, criá-la
if not os.path.exists('ler'):
    os.makedirs('ler')
   
    
def login(usuario, senha, banco, cpfs):

# Inicialize o navegador
    browser = Browser('chrome', headless=False)
    

# Acesse a página
    url = 'https://www.portaldoconsignado.com.br/'
    browser.visit(url)
    admin_element = browser.find_by_xpath('//*[@id="guias"]/div[2]/span/span')
    admin_element.click()
#recebe o valor do elemento e preenche o campo
    usuario_element = browser.find_by_xpath('//*[@id="username"]')
    usuario_element.fill(usuario)

#recebe o valor do elemento e preenche o campo
    senha_element = browser.find_by_xpath('//*[@id="password"]')
    senha_element.fill(senha)

    while True:
    # Localize o elemento do captcha e capture a imagem
        element = browser.find_by_xpath('//*[@id="captchaImg"]/div[1]')
        location = element._element.location
        size = element._element.size
        browser.driver.save_screenshot('ler/TELA_PORTAL.png')
        img = Image.open('ler/TELA_PORTAL.png')
        left = location['x']
        top = location['y']
        right = location['x'] + size['width']
        bottom = location['y'] + size['height']
        img = img.crop((left, top, right, bottom))
        img.save('ler/captcha.png')
    
    # Quebra o captcha e verifica se tem 6 caracteres
        texto_captcha = quebrar_captcha()
        logging.info(f"Meu texto do CAPTCHA foi: {texto_captcha}")

    
        if len(texto_captcha) == 6:
        # Preenche o captcha com o texto obtido
            campo_captcha = browser.find_by_xpath('//*[@id="captcha"]')
            campo_captcha.fill(texto_captcha)
        
        # Clique no botão para enviar o captcha preenchido
            botao = browser.find_by_xpath('//*[@id="idc"]')
            botao.click()
            time.sleep(2)

        #se esse elemento aparecer //*[@id="btRecolher"] clique nele
            botao_recolher = browser.find_by_xpath('//*[@id="btRecolher"]')
            if botao_recolher:
                botao_recolher.click()
           
                #botao select
            
                botoes = {
                    'Daycoval': '//*[@id="divlistaPerfil"]/fieldset/span/label[1]',
                    'Olé': '//*[@id="divlistaPerfil"]/fieldset/span/label[2]',
                    'Santander': '//*[@id="divlistaPerfil"]/fieldset/span/label[3]',}

# Clique no botão do banco selecionado
                botao_banco = browser.find_by_xpath(botoes[banco])
                botao_banco.click()


            #finalizar o botao select

            # eu quero apertar enter na pagina
                botao = browser.find_by_css('.botaoAcessar').first
                botao.click()
                time.sleep(1)
            # Localiza e clica no botão que exibe o menu
                botao = browser.find_by_text('Consulta de Margem').first
                botao.click()
                time.sleep(1)

                # Localiza e clica no link "Consulta de Margem" no menu expandido
                element = browser.find_by_xpath('//a[@class="nivel2"][@href="/consignatario/pesquisarMargem"]')
                element.click()
                
            
                
# adicionar cabeçalhos de coluna à planilha
                    # adicionar cabeçalhos de coluna à planilha, somente na primeira vez que a função for chamada
                
                if not os.path.isfile('ler/resultado.xlsx'):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['Nome', 'CPF', 'Órgão', 'Matrícula', 'MARGEM PARA EMPRESTIMO', 'MARGEM PARA CARTÃO', 'Consignações Compulsórias','simulacao'])
                    
                if 'ws' not in locals():
                    wb = load_workbook('ler/resultado.xlsx')
                    ws = wb.active
                for cpf in cpfs:
                    cpf_servidor = browser.find_by_xpath('//*[@id="cpfServidor"]')
                    cpf_servidor.fill(cpf)
                    time.sleep(1)
                    botao_pesquisar = browser.find_by_name('botaoPesquisar').first
                    botao_pesquisar.click()
                    time.sleep(1)
                    if browser.is_element_present_by_xpath('//*[@id="divEtapaError2"]'):
                        print(f"Erro ao processar CPF {cpf}. Indo para o próximo CPF...")
                        cpf_servidor = browser.find_by_xpath('//*[@id="cpfServidor"]')
                        
                        cpf_servidor.fill(cpf)
                        continue
                    else:
                        div_resultado_servidor = browser.find_by_id('divResultadoServidor').first
                        if div_resultado_servidor is None:
                            print(f"Erro ao processar CPF {cpf}. Indo para o próximo CPF...")
                            cpf_servidor = browser.find_by_xpath('//*[@id="cpfServidor"]')
                            
                            cpf_servidor.fill(cpf)
                            continue
                        try:
                            tabela = browser.find_by_xpath('//table[@id="tabelaMargem"]')[0]
                            if tabela is None:
                            
                                print(f"Erro ao processar CPF {cpf}. Indo para o próximo CPF...")
                        except:
                                cpf_servidor = browser.find_by_xpath('//*[@id="cpfServidor"]')
                                
                                cpf_servidor.fill(cpf)
                                continue
    # Resto do código...

                        
                    
                    cpf_element = div_resultado_servidor.find_by_xpath("//div[contains(text(), 'CPF')]/span").first
                    nome_element = div_resultado_servidor.find_by_xpath("//div[contains(text(), 'Nome')]/span").first
                    orgao_element = div_resultado_servidor.find_by_xpath("//div[contains(text(), 'Órgão')]/span").first
                    matricula_element = div_resultado_servidor.find_by_xpath("//div[contains(text(), 'Identificação')]/span").first
                    cpf = cpf_element.text
                    nome = nome_element.text
                    orgao = orgao_element.text
                    matricula = matricula_element.text

                    
                    tabela = browser.find_by_xpath('//table[@id="tabelaMargem"]')[0]
                    linhas = tabela.find_by_xpath('//tr[@id="linha"]')
                    ultimas_linhas = linhas[-3:]
                    produtos = []
                    valores = []
                    for linha in ultimas_linhas:
                        produto = linha.find_by_xpath('.//td[1]/span')[0].text
                        valor = linha.find_by_xpath('.//td[2]/span')[0].text
                        produtos.append(produto)
                        valores.append(valor)
                        

                        
                        

                        
                    ws.append([nome, cpf, orgao, matricula, *valores ])
                    
                    wb.save('Margens_Disponivel.xlsx')
                    workbook = openpyxl.load_workbook('Margens_Disponivel.xlsx')
                    worksheet = workbook.active
                    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):
                        for cell in row:
                            if cell.value is not None:
                                cell.offset(column=2).value = f"=F{cell.row}/0.0480693"
                                for cell in ws['H']:
                                    cell.number_format = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-;_-@_-'


                                workbook.save('Margens_Disponivel.xlsx')
                                print('Planilha atualizada com sucesso!')

                   




                    


 
                


                
        # Verifique se a mensagem de erro apareceu
            mensagem_erro = browser.find_by_xpath('//*[@id="id14"]/ul/li/span')
            if mensagem_erro:
                logging.info("Os caracteres digitados não correspondem à imagem. Refazendo processo...")
            #se aparecer  digite a senha novamente e continue o loop
                senha_element = browser.find_by_xpath('//*[@id="password"]')
                senha_element.fill(senha)
                continue

        # Se não houve erro, saia do loop
            break
    
    # Se o captcha tiver menos de 6 caracteres, clique no botão para atualizar a imagem
        botao_atualizar = browser.find_by_xpath('//*[@id="ida"]')
        botao_atualizar.click()
        time.sleep(0.5)
 


    
    browser.quit()